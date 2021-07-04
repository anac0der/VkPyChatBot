import datetime
import openpyxl


def CurrentWeekday():
    curr_date = datetime.datetime.now()
    return curr_date.isoweekday()


def ToSec(h, mins, sec=0):
    return h * 3600 + mins * 60 + sec


def LessonNum(time): #время в секундах от начала дня
    if time >= 55800:
        return None
    if time < 30600:
        return 1
    elif time < 34200:
        return 2
    elif time < 37800:
        return 3
    elif time < 41700:
        return 4
    elif time < 45000:
        return 5
    elif time < 48900:
        return 6
    elif time < 52200:
        return 7
    elif time < 55800:
        return 8


def ClassColumn(class_):
    wb = openpyxl.load_workbook(filename="schedule.xlsx")
    schedule = wb['schedule']
    for i in range(50):
        cell = schedule.cell(row=1, column=i+2)
        if cell.value == class_:
            return i+2


def CurrentLesson(class_):
    time = datetime.datetime.now()
    curr_time = ToSec(time.timetuple()[3], time.timetuple()[4], time.timetuple()[5])
    if CurrentWeekday() > 5 or not LessonNum(curr_time):
        return None
    else:
        wb = openpyxl.load_workbook(filename="schedule.xlsx")
        schedule = wb['schedule']

        row = LessonNum(curr_time) + (CurrentWeekday() - 1) * 10 + 1
        lesson = schedule.cell(row=row, column=ClassColumn(class_))

        return lesson.value


'''#'тест:
time = datetime.datetime.now() 
curr_time = ToSec(time.timetuple()[3], time.timetuple()[4], time.timetuple()[5])
print(CurrentLesson('11е'))
print(LessonNum(curr_time))
row = LessonNum(curr_time) + (CurrentWeekday()) * 10
print(row)
print(ClassColumn('11е'))'''